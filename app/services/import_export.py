"""Import / Export Engine.

Doc 03 §26/§27:
    - La plantilla se genera desde la configuración, no es una planilla fija.
    - La importación es atómica: validar todo primero, después commitear todo
      o nada. Nunca se escribe parcialmente.
    - Un error informa fila, columna, valor, error y corrección esperada.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..domain.config import Configuration, SalesMode
from ..domain.inputs import Concept, InputSet, InputSource, InputValue
from ..domain.periods import Period
from .budget import BudgetError, BudgetVersion

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LOCKED_FILL = PatternFill("solid", fgColor="EDEDED")


@dataclass
class ImportError_:
    row: int
    column: str
    value: str
    error: str
    expected: str

    def as_dict(self) -> dict:
        return {"row": self.row, "column": self.column, "value": self.value,
                "error": self.error, "expected": self.expected}


@dataclass
class ImportResult:
    status: str                       # COMMITTED | REJECTED
    imported: int = 0
    errors: list[ImportError_] = field(default_factory=list)

    def as_dict(self) -> dict:
        return {"status": self.status, "imported": self.imported,
                "errors": [e.as_dict() for e in self.errors]}


# ==========================================================================
# Plantilla de ventas
# ==========================================================================
def sales_template(version: BudgetVersion, operation_id: str) -> bytes:
    """Genera la planilla que le corresponde a esa operación y a nadie más."""
    cfg = version.configuration
    op = cfg.operation(operation_id)
    unit = cfg.unit(op.business_unit_id)
    branch = cfg.branch(op.branch_id)
    fy = cfg.fiscal_year

    wb = Workbook()
    ws = wb.active
    ws.title = "Carga"
    # La modalidad es del producto: en la misma planilla puede haber productos
    # que se cargan por cantidad y otros por monto.
    headers = ["UNIDAD", "SUCURSAL", "PRODUCTO", "CODIGO", "MODALIDAD", "FRECUENCIA",
               "PERIODO", "MONEDA", "VALOR"]
    ws.append(headers)
    for i, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    row = 2
    for product in unit.products:
        for head, bucket in fy.iter_buckets(product.sales_frequency):
            if not any(cfg.is_active(op, p) and cfg.is_active(unit, p)
                       and cfg.is_active(branch, p) for p in bucket):
                continue
            unit_based = product.sales_mode is SalesMode.UNIT_BASED
            ws.append([unit.name, branch.name, product.name, product.code,
                       "Cantidad" if unit_based else "Monto",
                       product.sales_frequency.value, head.code,
                       "unidades" if unit_based else product.currency, None])
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = LOCKED_FILL
            row += 1

    widths = [22, 22, 26, 10, 12, 14, 12, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    info = wb.create_sheet("Instrucciones")
    info["A20"] = "OPERACION"
    info["B20"] = operation_id
    info["A1"] = "Instrucciones de carga"
    info["A1"].font = Font(bold=True, size=13)
    lines = [
        "La estructura de esta planilla la define la configuración aprobada; no se puede modificar.",
        "Cargue únicamente la columna VALOR: cantidad o monto según diga MODALIDAD.",
        "0 es un valor válido. Una celda vacía es un error si el campo es obligatorio.",
        "Si esta operación no vende un producto, cargue 0.",
        "La importación es atómica: un solo error rechaza toda la planilla.",
        "El precio y el margen los define la configuración; no se cargan acá.",
    ]
    for i, t in enumerate(lines, start=3):
        info[f"A{i}"] = f"• {t}"
    info.column_dimensions["A"].width = 110

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_sales_import(version: BudgetVersion, data: bytes, operation_id: str,
                       actor: str) -> tuple[ImportResult, list[InputValue]]:
    cfg = version.configuration
    op = cfg.operation(operation_id)
    unit = cfg.unit(op.business_unit_id)
    fy = cfg.fiscal_year
    errors: list[ImportError_] = []
    parsed: list[InputValue] = []

    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception:
        return ImportResult("REJECTED", 0, [ImportError_(
            0, "archivo", "", "El archivo no es una planilla Excel válida",
            "Descargá la plantilla desde el sistema, completala y subila sin convertirla.")]), []
    ws = wb["Carga"] if "Carga" in wb.sheetnames else wb.active

    # La plantilla lleva marcada su operación: sin esto, la planilla de una
    # sucursal entraba en la tarea de otra y le pisaba las ventas.
    marca = wb["Instrucciones"]["B20"].value if "Instrucciones" in wb.sheetnames else None
    if marca and str(marca).strip() != operation_id:
        return ImportResult("REJECTED", 0, [ImportError_(
            0, "archivo", str(marca),
            f"Esta planilla es de {cfg.operation_label(str(marca).strip())}",
            f"Subí la planilla de {cfg.operation_label(operation_id)}, "
            "o descargala de nuevo desde esta tarea.")]), []
    headers = [str(c.value or "").strip().upper() for c in ws[1]]
    try:
        i_code = headers.index("CODIGO")
        i_period = headers.index("PERIODO")
        i_ccy = headers.index("MONEDA")
        i_value = (headers.index("VALOR") if "VALOR" in headers
                   else headers.index("CANTIDAD") if "CANTIDAD" in headers
                   else headers.index("MONTO"))
    except ValueError:
        return ImportResult("REJECTED", 0, [ImportError_(
            1, "encabezados", ",".join(headers), "La planilla no tiene la estructura esperada",
            "Descargue la plantilla desde el sistema y no modifique los encabezados.")]), []

    by_code = {p.code: p for p in unit.products}

    for r in range(2, ws.max_row + 1):
        code = ws.cell(row=r, column=i_code + 1).value
        if code is None:
            continue
        code = str(code).strip()
        raw = ws.cell(row=r, column=i_value + 1).value
        period_raw = str(ws.cell(row=r, column=i_period + 1).value or "").strip()
        ccy = str(ws.cell(row=r, column=i_ccy + 1).value or "").strip().upper()

        product = by_code.get(code)
        if product is None:
            errors.append(ImportError_(r, "CODIGO", code, "Producto inexistente en la unidad",
                                       f"Uno de: {', '.join(sorted(by_code))}"))
            continue
        unit_based = product.sales_mode is SalesMode.UNIT_BASED
        if raw is None or str(raw).strip() == "":
            errors.append(ImportError_(
                r, "VALOR", "", "Celda vacía",
                "Un número. Si no corresponde, cargue 0 (0 es válido; vacío no)."))
            continue
        try:
            value = Decimal(str(raw).replace(",", "."))
        except (InvalidOperation, ValueError):
            errors.append(ImportError_(r, "VALOR", str(raw),
                                       "No es un número", "Un valor numérico, sin texto ni símbolos"))
            continue
        if value < 0:
            errors.append(ImportError_(r, "VALOR", str(raw),
                                       "Valor negativo", "Un valor mayor o igual a 0"))
            continue
        try:
            period = Period.parse(period_raw)
            head = fy.bucket_head(period, product.sales_frequency)
        except Exception:
            errors.append(ImportError_(r, "PERIODO", period_raw, "Período inválido",
                                       "Formato AAAA-MM dentro del ejercicio"))
            continue
        if period != head:
            errors.append(ImportError_(
                r, "PERIODO", period_raw,
                f"El producto se carga con frecuencia {product.sales_frequency.value}",
                f"El período de carga debe ser {head.code}"))
            continue

        parsed.append(InputValue(
            concept=Concept.SALES_QTY if unit_based else Concept.SALES_AMOUNT,
            period=period.code, value=value,
            currency=None if unit_based else (ccy if ccy and ccy != "UNIDADES"
                                              else product.currency),
            operation_id=operation_id, business_unit_id=unit.id,
            branch_id=op.branch_id, product_id=product.id,
            source=InputSource.IMPORT, loaded_by=actor,
        ))

    if errors:
        return ImportResult("REJECTED", 0, errors), []
    esperadas = sum(1 for product in unit.products
                    for head, bucket in fy.iter_buckets(product.sales_frequency)
                    if any(cfg.is_active(op, p) and cfg.is_active(unit, p)
                           and cfg.is_active(cfg.branch(op.branch_id), p) for p in bucket))
    if len(parsed) < esperadas:
        errors.append(ImportError_(
            0, "archivo", str(len(parsed)),
            f"La planilla trae {len(parsed)} filas y la operación tiene {esperadas}",
            "No borres filas: si un producto no se vende en un período, cargá 0."))
        return ImportResult("REJECTED", 0, errors), []
    return ImportResult("COMMITTED", len(parsed)), parsed


def commit_import(service, actor: str, version: BudgetVersion,
                  parsed: list[InputValue]) -> ImportResult:
    """Commit atómico: si algo falla, no queda nada escrito."""
    version.assert_mutable()
    snapshot = version.inputs.model_copy(deep=True)
    estados = {t.id: (t.status, len(t.history)) for t in version.tasks.values()}
    eventos = len(service.audit.events)
    try:
        for iv in parsed:
            service.submit_input(actor, version, iv, capability="budget.sales.load")
    except Exception as exc:
        # "O todo o nada" vale también para las aprobaciones que la carga
        # invalidó y para los eventos de auditoría de cambios que no ocurrieron.
        version.inputs = snapshot
        version.invalidate()
        for t in version.tasks.values():
            if t.id in estados:
                t.status, hasta = estados[t.id]
                del t.history[hasta:]
        del service.audit.events[eventos:]
        raise BudgetError("IMPORT_VALIDATION_FAILED", str(exc))
    return ImportResult("COMMITTED", len(parsed))
