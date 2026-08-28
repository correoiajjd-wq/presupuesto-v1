"""Tests de las reglas de negocio del spec V1.

Cada test cita la regla que verifica. Ejecutar con:
    PYTHONPATH=. python3 -m unittest discover -s tests -v
"""
from __future__ import annotations

import unittest
from datetime import date
from decimal import Decimal

from app.domain.config import (
    AllocationMode, ConfigStatus, ConfigurationError, ExpenseDefinition, ExpenseTarget,
    ExpenseTargetType, MarginFormula, Role, SalesMode,
)
from app.domain.engine import (
    FY, BudgetEngine, scope_br, scope_bu, scope_co, scope_op, scope_stock,
)
from app.domain.graph import nk
from app.domain.inputs import ChangeType, Concept, InputSet, InputValue
from app.domain.money import FXTable, Money
from app.domain.periods import FiscalYear, Frequency, Period, spread
from app.domain.ratios import RATIO_CATALOG
from app.domain.validation import (
    Severity, evaluate_objectives, missing_required_inputs, validate_balance, validate_inputs,
)
from app.services.budget import BudgetError, TaskStatus, VersionStatus
from app.services.scenarios import ScenarioAdjustment, apply_overlay, compare, run_scenario
from app.services.budget import Scenario
from seed.demo import D, bootstrap, build_configuration, build_fx, build_inputs


def val(values, metric, scope, period=FY):
    return values.get(nk(metric, scope, period))


class TestPeriods(unittest.TestCase):
    def test_fiscal_year_no_calendario(self):
        """Doc 01 §3: el ejercicio puede empezar en cualquier fecha."""
        fy = FiscalYear(date(2027, 4, 1), date(2028, 3, 31))
        self.assertEqual(len(fy.periods), 12)
        self.assertEqual(fy.periods[0].code, "2027-04")
        self.assertEqual(fy.periods[-1].code, "2028-03")

    def test_balance_inicial_es_dia_anterior(self):
        """Doc 02 §33."""
        fy = FiscalYear(date(2027, 1, 1), date(2027, 12, 31))
        self.assertEqual(fy.opening_balance_date, date(2026, 12, 31))

    def test_distribucion_equitativa_sin_perder_centavos(self):
        """Doc 01 §23: USD 120.000 anual -> 10.000/mes."""
        fy = FiscalYear(date(2027, 1, 1), date(2027, 12, 31))
        out = spread(D(120_000), fy.periods)
        self.assertEqual(len(out), 12)
        self.assertTrue(all(v == D(10_000) for v in out.values()))
        # un monto no divisible no puede perder ni ganar centavos
        out = spread(D(100), fy.periods)
        self.assertEqual(sum(out.values()), D(100))


class TestFX(unittest.TestCase):
    def test_moneda_puente(self):
        """Doc 01 §4: ARS -> USD -> UYU sin cargar el par ARS/UYU."""
        fx = FXTable("USD", ["USD", "UYU", "ARS"])
        p = Period(2027, 1)
        fx.add_flat("UYU", date(2027, 1, 1), date(2027, 1, 31), D("0.025"))
        fx.add_flat("ARS", date(2027, 1, 1), date(2027, 1, 31), D("0.001"))
        out = fx.convert(Money.of(1_000_000, "ARS"), "UYU", p)
        self.assertEqual(out.currency, "UYU")
        self.assertEqual(out.amount, D(40_000))   # 1M ARS = 1.000 USD = 40.000 UYU

    def test_moneda_no_habilitada_se_rechaza(self):
        fx = FXTable("USD", ["USD"])
        with self.assertRaises(Exception):
            fx.add("BRL", date(2027, 1, 1), D("0.2"))

    def test_faltan_dias_de_tc(self):
        """Doc 02 §30: el TC se carga para cada día del ejercicio."""
        fx = FXTable("USD", ["USD", "UYU"])
        fx.add("UYU", date(2027, 1, 1), D("0.025"))
        gaps = fx.coverage_gaps(["UYU"], date(2027, 1, 1), date(2027, 1, 31))
        self.assertIn("faltan 30", gaps[0])


class TestConfiguration(unittest.TestCase):
    def test_otros_obligatorio_por_familia(self):
        """Doc 02 §8/§9: el 'Otros' se controla por familia, no por unidad."""
        cfg = build_configuration()
        unidad = cfg.business_units[0]
        self.assertEqual(unidad.missing_other_products(), [])
        # sacarle el "Otros" a una sola familia ya deja la configuración incompleta
        unidad.products = [p for p in unidad.products if p.code != "XXACC"]
        faltantes = [f.name for f in unidad.missing_other_products()]
        self.assertEqual(faltantes, ["Accesorios"])
        self.assertTrue(any("Accesorios" in e for e in cfg.validate_structure()))

    def test_dos_otros_en_la_misma_familia_se_rechaza(self):
        cfg = build_configuration()
        unidad = cfg.business_units[0]
        data = unidad.model_dump()
        data["products"].append({**data["products"][-1], "id": "P-900", "code": "ZZ",
                                 "family_id": "FAM-ACC", "is_other": True})
        with self.assertRaises(Exception) as ctx:
            type(unidad)(**data)
        self.assertIn("DUPLICATE_OTHER_PRODUCT", str(ctx.exception))

    def test_distribucion_debe_sumar_100(self):
        """Doc 02 §21."""
        with self.assertRaises(Exception) as ctx:
            ExpenseDefinition(
                id="X", name="X", currency="USD",
                allocation_mode=AllocationMode.PERCENTAGE,
                targets=[ExpenseTarget(target_type=ExpenseTargetType.BUSINESS_UNIT,
                                       target_id="BU-01", percentage=D("0.5"))],
            )
        self.assertIn("INVALID_ALLOCATION", str(ctx.exception))

    def test_gasto_sin_destino_se_rechaza(self):
        with self.assertRaises(Exception) as ctx:
            ExpenseDefinition(id="X", name="Internet", currency="USD", targets=[])
        self.assertIn("INVALID_EXPENSE", str(ctx.exception))

    def test_ratio_arrastra_dependencias(self):
        """Doc 02 §37: elegir un ratio de stock exige configurar stock."""
        cfg = build_configuration()
        cfg.inventory.enabled = False
        pending = cfg.missing_modules_for_ratios()
        self.assertTrue(any("Stock" in m for m in pending))

    def test_los_porcentajes_se_escriben_como_fraccion(self):
        """Un 25 donde va 0.25 multiplicaba la nómina por cien sin que nada lo
        frenara: sólo el margen de producto estaba validado."""
        cfg = build_configuration()
        cfg.business_units[1].products[0].commission_rate = D(5)
        cfg.payroll.percentage_concepts[0].percentage = D(12)
        cfg.payroll.increase_rules[0].percentage = D(8)
        codigos = [e.split(":")[0] for e in cfg.validate_structure()]
        self.assertEqual(codigos.count("INVALID_PERCENTAGE"), 3)

    def test_obligatoriedad_surge_del_modelo(self):
        """Doc 02 §42: no hay lista rígida de datos obligatorios."""
        cfg = build_configuration()
        required = cfg.required_concepts()
        self.assertIn("SALES", required)
        self.assertIn("OPENING_STOCK", required)
        cfg.inventory.enabled = False
        cfg.ratios = [r for r in cfg.ratios if "STOCK" not in r.ratio_code]
        self.assertNotIn("OPENING_STOCK", cfg.required_concepts())


class TestSalesAndMargin(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.values = self.version.calculate()

    def test_modalidad_unidades(self):
        """Doc 02 §10: el gerente carga cantidad; el sistema calcula ventas y costo."""
        v = val(self.values, "SALES", "OP:OP-01#P:P-001", "2027-01")
        self.assertEqual(v, D(2500) * D(100))
        c = val(self.values, "COGS", "OP:OP-01#P:P-001", "2027-01")
        self.assertEqual(c, v * (Decimal(1) - D("0.30")))

    def test_modalidad_monto_convierte_moneda(self):
        """BU-02 carga en UYU; el sistema presenta en USD."""
        v = val(self.values, "SALES", "OP:OP-03#P:P-101", "2027-01")
        self.assertEqual(v, D(12_000_000) * D("0.025"))

    def test_margen_markup_sobre_costo(self):
        """Doc 02 §11: la fórmula de margen es configuración, y es del producto."""
        sales = val(self.values, "SALES", "OP:OP-01#P:P-002", "2027-01")
        cogs = val(self.values, "COGS", "OP:OP-01#P:P-002", "2027-01")
        self.assertAlmostEqual(float(cogs), float(sales / (Decimal(1) + D("0.35"))), places=6)

    def test_producto_sin_costo(self):
        """Un intangible: el precio de venta es todo margen."""
        sales = val(self.values, "SALES", "OP:OP-03#P:P-102", "2027-01")
        cogs = val(self.values, "COGS", "OP:OP-03#P:P-102", "2027-01")
        self.assertGreater(sales, 0)
        self.assertEqual(cogs, D(0))

    def test_modalidad_es_del_producto_no_de_la_unidad(self):
        """La misma unidad puede tener productos por unidades y por monto."""
        modos = {p.sales_mode.value for u in self.version.configuration.business_units
                 for p in u.products}
        self.assertEqual(modos, {"UNIT_BASED", "AMOUNT_BASED"})
        formulas = {p.margin_formula.value for u in self.version.configuration.business_units
                    for p in u.products}
        self.assertIn("MARKUP_ON_COST", formulas)
        self.assertIn("NO_COST", formulas)

    def test_frecuencia_trimestral_se_distribuye(self):
        """Doc 02 §13: si se carga en frecuencia mayor, se distribuye equitativamente."""
        q1 = [val(self.values, "SALES", "OP:OP-01#P:P-002", f"2027-0{m}") for m in (1, 2, 3)]
        # partes iguales salvo el ajuste de redondeo, que va al último mes
        self.assertLessEqual(max(q1) - min(q1), D("0.01") * D(250))
        self.assertEqual(sum(q1), D(400) * D(250))   # no se pierde ni se inventa valor

    def test_sucursal_que_abre_en_junio(self):
        """Doc 02 §7: el sistema respeta las fechas de vigencia."""
        for m in ("2027-01", "2027-05"):
            self.assertEqual(val(self.values, "SALES", scope_br("BR-02"), m), D(0))
        self.assertGreater(val(self.values, "SALES", scope_br("BR-02"), "2027-06"), D(0))

    def test_trimestre_parcial_no_se_reparte_a_meses_cerrados(self):
        """El trimestre abr-jun de una sucursal que abre en junio va entero a junio."""
        abr = val(self.values, "SALES", "OP:OP-02#P:P-002", "2027-04")
        jun = val(self.values, "SALES", "OP:OP-02#P:P-002", "2027-06")
        self.assertEqual(abr, D(0))
        self.assertEqual(jun, D(150) * D(250))


class TestPayroll(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.values = self.version.calculate()
        self.cfg = self.version.configuration

    def test_un_aumento_alcanza_a_quien_ya_estaba(self):
        """Doc 01 §17: el alta de febrero cobra el aumento de marzo porque ya
        estaba; el alta de junio no, porque entró después."""
        e = BudgetEngine(self.cfg, self.version.fx, self.version.inputs)
        self.assertEqual(e._increase_factor(date(2027, 2, 1), Period(2027, 3)), D("1.05"))
        self.assertEqual(e._increase_factor(date(2027, 6, 1), Period(2027, 7)), D(1))
        self.assertEqual(e._increase_factor(date(2027, 6, 1), Period(2027, 8)), D("1.04"))
        self.assertEqual(e._increase_factor(date(2027, 1, 1), Period(2027, 12)),
                         D("1.05") * D("1.04"))

    def test_el_alta_arrastra_sus_propios_aumentos(self):
        """El alta de Salto entra en junio: no cobra el aumento de marzo."""
        cargas = D("1.17")
        junio = val(self.values, "PAYROLL_BASE", "CC:CC-102", "2027-06")
        self.assertEqual(junio, (D(6_400) * D("1.05") + D(4_200)) * cargas)
        agosto = val(self.values, "PAYROLL_BASE", "CC:CC-102", "2027-08")
        self.assertEqual(agosto,
                         (D(6_400) * D("1.05") * D("1.04") + D(4_200) * D("1.04")) * cargas)

    def test_la_baja_paga_su_mes_y_descuenta_desde_el_siguiente(self):
        """La baja es del 1/9: septiembre se paga igual y el descuento arranca
        en octubre, arrastrando los aumentos que esa persona ya tenía."""
        cargas = D("1.17")
        agosto = val(self.values, "PAYROLL_BASE", "CC:CC-101", "2027-08")
        septiembre = val(self.values, "PAYROLL_BASE", "CC:CC-101", "2027-09")
        octubre = val(self.values, "PAYROLL_BASE", "CC:CC-101", "2027-10")
        self.assertEqual(septiembre, agosto)
        self.assertEqual(octubre, septiembre - D(2_100) * D("1.05") * D("1.04") * cargas)
        # y las personas siguen el mismo criterio que la plata
        self.assertEqual(val(self.values, "HEADCOUNT", "CC:CC-101", "2027-09"), D(10))
        self.assertEqual(val(self.values, "HEADCOUNT", "CC:CC-101", "2027-10"), D(9))

    def test_la_cantidad_autorizada_reajusta_el_importe_sola(self):
        """El gerente pide 2, Nómina valoriza, la revisión autoriza 1: el importe
        se recalcula con una regla de tres, sin volver a pedirle nada a Nómina."""
        mov = next(iv for iv in self.version.inputs.values
                   if iv.concept is Concept.HEADCOUNT_CHANGE and iv.movement_id == "MOV-01")
        antes = val(self.values, "PAYROLL_BASE", "CC:CC-101", "2027-02")
        mov.value = D(1)
        self.version.invalidate()
        despues = self.version.calculate()
        self.assertEqual(antes - val(despues, "PAYROLL_BASE", "CC:CC-101", "2027-02"),
                         D(2_500) * D("1.17"))
        self.assertEqual(val(despues, "HEADCOUNT", "CC:CC-101", "2027-02"), D(9))
        # y no quedó nada pendiente de valorizar
        self.assertFalse([f for f in missing_required_inputs(self.cfg, self.version.inputs)
                          if "valorizar" in f.message])

    def test_el_ajuste_mueve_plata_y_no_personas(self):
        """Un ascenso: sube el nominal desde julio sin tocar la dotación."""
        cargas = D("1.17")
        junio = val(self.values, "PAYROLL_BASE", "CC:CC-01", "2027-06")
        julio = val(self.values, "PAYROLL_BASE", "CC:CC-01", "2027-07")
        self.assertEqual(julio - junio, D(900) * cargas)
        personas = {val(self.values, "HEADCOUNT", "CC:CC-01", p.code)
                    for p in self.cfg.periods}
        self.assertEqual(personas, {D(3)})

    def test_dotacion_altas_y_bajas(self):
        """Doc 02 §54: dotación inicial + altas - bajas = dotación final."""
        ene = val(self.values, "HEADCOUNT", scope_op("OP-01"), "2027-01")
        feb = val(self.values, "HEADCOUNT", scope_op("OP-01"), "2027-02")
        oct = val(self.values, "HEADCOUNT", scope_op("OP-01"), "2027-10")
        self.assertEqual(ene, D(8))          # foto inicial que cargó Nómina
        self.assertEqual(feb, D(10))         # +2 altas
        self.assertEqual(oct, D(9))          # -1 baja del 1/9, efectiva en octubre
        # La sucursal Montevideo aloja además la operación de Servicios: su
        # dotación es la suma de las dos, no la de una sola.
        self.assertEqual(val(self.values, "HEADCOUNT", scope_br("BR-01"), "2027-01"),
                         ene + val(self.values, "HEADCOUNT", scope_op("OP-04"), "2027-01"))

    def test_costo_laboral_incluye_cargas(self):
        """Doc 01 §18: los conceptos porcentuales se aplican automáticamente.

        El costo sale del nominal que Nómina cargó contra el centro de costo de
        esa operación, no de la dotación.
        """
        enero = val(self.values, "PAYROLL_BASE", scope_op("OP-01"), "2027-01")
        self.assertEqual(enero, D(20_500) * D("1.17"))

    def test_ser_responsable_de_un_centro_habilita_a_pedir_gente(self):
        """El CFO designa al responsable al crear el centro de costo; de ahí sale
        el permiso, sin tener que agregarle roles a la tabla de capacidades."""
        auth = self.service.auth
        ana = self.service.user("u.admin")          # ADMIN_AREA, responsable de CC-01
        self.assertNotIn("budget.headcount.load", auth.capabilities(ana))
        auth.check(ana, "budget.headcount.load", "CC:CC-01", self.cfg)
        # pero no en un centro que no es suyo
        with self.assertRaises(BudgetError) as ctx:
            auth.check(ana, "budget.headcount.load", "CC:CC-101", self.cfg)
        self.assertEqual(ctx.exception.code, "UNAUTHORIZED")
        # y sigue sin poder valorizar: eso es de Nómina
        with self.assertRaises(BudgetError):
            auth.check(ana, "budget.payroll.load", "CC:CC-01", self.cfg)

    def test_una_solicitud_sin_valorizar_bloquea(self):
        """El área pide, Nómina valoriza. Hasta que no lo haga, no se aprueba."""
        self.service.submit_input("u.br03", self.version, InputValue(
            concept=Concept.HEADCOUNT_CHANGE, value=D(10), change_type=ChangeType.HIRED,
            effective_date=date(2027, 6, 1), cost_center_id="CC-103",
            comment="diez vendedores", movement_id="MOV-99"),
            "budget.headcount.load")
        faltantes = missing_required_inputs(self.cfg, self.version.inputs)
        self.assertTrue(any("valorizar" in f.message for f in faltantes))
        self.assertTrue(all(f.blocking for f in faltantes))
        # la solicitud ya mueve las personas, pero todavía no la plata
        despues = self.version.calculate(force=True)
        self.assertEqual(val(despues, "HEADCOUNT", "CC:CC-103", "2027-06"), D(14))
        self.assertEqual(val(despues, "PAYROLL_BASE", "CC:CC-103", "2027-06"),
                         val(self.values, "PAYROLL_BASE", "CC:CC-103", "2027-06"))

    def test_la_foto_inicial_no_admite_otro_ambito_ni_periodo(self):
        """El motor agrupa la nómina por centro de costo. Si el mismo dato entra
        además con otro ámbito o con período, se suma dos veces."""
        for iv in (InputValue(concept=Concept.NOMINAL_SALARY, value=D(20_500), currency="USD",
                              cost_center_id="CC-101", operation_id="OP-01"),
                   InputValue(concept=Concept.NOMINAL_SALARY, value=D(20_500), currency="USD",
                              cost_center_id="CC-101", period="2027-01")):
            with self.assertRaises(BudgetError):
                self.service.submit_input("u.payroll", self.version, iv, "budget.payroll.load")
        self.assertEqual(self.version.calculate()[nk("PAYROLL_BASE", "CC:CC-101", "2027-01")],
                         D(20_500) * D("1.17"))

    def test_la_foto_inicial_la_carga_nomina_por_centro_de_costo(self):
        cargados = {iv.cost_center_id for iv in self.version.inputs.values
                    if iv.concept is Concept.INITIAL_HEADCOUNT}
        self.assertEqual(cargados, {cc.id for cc, _k, _l in self.cfg.cost_centers()})

    def test_comisiones_se_calculan_desde_ventas(self):
        """Doc 01 §19: Ventas -> Nómina -> comisión.

        La tasa es de cada producto: dentro de la misma sucursal, unos
        comisionan, otros comisionan distinto y otros no comisionan.
        """
        unidad = self.cfg.unit("BU-02")
        esperado = D(0)
        con_comision = 0
        for prod in unidad.products:
            ventas = val(self.values, "SALES", f"OP:OP-03#P:{prod.id}", "2027-01")
            if prod.commission_rate:
                esperado += ventas * prod.commission_rate
                con_comision += 1
        comision = val(self.values, "COMMISSION", scope_op("OP-03"), "2027-01")
        self.assertEqual(comision, esperado)
        self.assertGreater(con_comision, 1)          # más de una tasa distinta
        self.assertLess(con_comision, len(unidad.products))   # y alguno sin comisión

    def test_un_producto_sin_comision_no_aporta(self):
        otros = self.cfg.unit("BU-02").product("P-199")
        self.assertIsNone(otros.commission_rate)
        ventas_otros = val(self.values, "SALES", "OP:OP-03#P:P-199", "2027-01")
        self.assertGreater(ventas_otros, 0)
        # la comisión total no incluye esas ventas
        total_ventas = val(self.values, "SALES", scope_op("OP-03"), "2027-01")
        comision = val(self.values, "COMMISSION", scope_op("OP-03"), "2027-01")
        self.assertLess(comision, total_ventas * D("0.05"))

    def test_la_comision_no_aplica_a_una_unidad_sin_tasas(self):
        """Repuestos no tiene productos con comisión: no genera comisión."""
        self.assertEqual(val(self.values, "COMMISSION", scope_op("OP-01"), "2027-01"), D(0))

    def test_nomina_de_soporte_no_esta_en_ebitda_de_la_unidad(self):
        pay_bu = sum(val(self.values, "PAYROLL", scope_bu(u.id)) for u in self.cfg.business_units)
        pay_co = val(self.values, "PAYROLL", scope_co())
        self.assertGreater(pay_co, pay_bu)


class TestExpenses(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.values = self.version.calculate()
        self.cfg = self.version.configuration

    def test_gasto_distribuido_por_porcentaje(self):
        """Doc 02 §21: el área carga el total, el sistema distribuye."""
        total = D(48_000)
        bu01 = val(self.values, "EXPENSES_UNIT_LEVEL", scope_bu("BU-01"))
        self.assertGreaterEqual(bu01, total * D("0.60") - D("0.01"))

    def test_un_gasto_va_a_varios_destinos(self):
        """Internet existe en las tres sucursales y en administración."""
        internet = self.cfg.expense("EXP-02")
        destinos = {t.scope_key for t in internet.targets}
        self.assertEqual(destinos, {"BR:BR-01", "BR:BR-02", "BR:BR-03", "CC:CC-01"})
        # cada destino tiene su propio importe
        self.assertEqual(val(self.values, "EXPENSE_INPUT", "EXP:EXP-02@BR:BR-01", "2027-01"),
                         D(600))
        self.assertEqual(val(self.values, "EXPENSE_INPUT", "EXP:EXP-02@BR:BR-03", "2027-01"),
                         D(500))

    def test_donde_no_corresponde_se_carga_cero(self):
        """Salto es propia: el alquiler existe como concepto pero vale 0."""
        self.assertEqual(val(self.values, "EXPENSE_INPUT", "EXP:EXP-01@BR:BR-02", "2027-01"),
                         D(0))
        self.assertGreater(val(self.values, "EXPENSE_INPUT", "EXP:EXP-01@BR:BR-01", "2027-01"),
                           D(0))

    def test_gasto_de_unidad_se_distribuye_proporcional_a_ventas(self):
        """Doc 02 §22: proporcional al volumen total de ventas.

        El marketing de Repuestos se reparte entre las operaciones de Repuestos
        —Montevideo y Salto— y no toca a la operación de Servicios que vive en
        la misma sucursal de Montevideo.
        """
        s1 = val(self.values, "SALES", scope_op("OP-01"))
        s2 = val(self.values, "SALES", scope_op("OP-02"))
        e1 = val(self.values, "EXPENSES", scope_op("OP-01"))
        e2 = val(self.values, "EXPENSES", scope_op("OP-02"))
        marketing = D(120_000)
        self.assertAlmostEqual(float(e1), float(marketing * s1 / (s1 + s2)), places=2)
        self.assertAlmostEqual(float(e2), float(marketing * s2 / (s1 + s2)), places=2)
        self.assertGreater(e1, e2)
        # Servicios Montevideo no recibe nada del marketing de Repuestos
        self.assertEqual(val(self.values, "EXPENSES", scope_op("OP-04")), D(0))

    def test_corporativos_se_muestran_debajo_del_ebitda(self):
        """Doc 02 §53: distinguir resultado propio del impacto corporativo."""
        eb_bu = val(self.values, "EBITDA", scope_bu("BU-01"))
        after = val(self.values, "RESULT_AFTER_ALLOCATION", scope_bu("BU-01"))
        alloc = val(self.values, "ALLOCATED_EXPENSES", scope_bu("BU-01"))
        self.assertAlmostEqual(float(after), float(eb_bu - alloc), places=4)
        self.assertGreater(alloc, D(0))
        # Lo que se le asigna a la unidad es lo corporativo y lo de las sucursales
        # donde opera; lo suyo propio ya está dentro de su EBITDA.
        por_operacion = sum(
            val(self.values, m, scope_op(o.id))
            for o in self.cfg.unit_operations("BU-01")
            for m in ("ALLOC_FROM_COMPANY", "ALLOC_FROM_BRANCH"))
        self.assertAlmostEqual(float(alloc), float(por_operacion), places=4)

    def test_la_asignacion_corporativa_cierra_contra_el_ebitda_de_la_empresa(self):
        suma = sum(val(self.values, "RESULT_AFTER_ALLOCATION", scope_bu(u.id))
                   for u in self.cfg.business_units)
        ebitda_co = val(self.values, "EBITDA", scope_co())
        self.assertAlmostEqual(float(suma), float(ebitda_co), places=4)


class TestInventory(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.values = self.version.calculate()

    def test_stock_final_es_calculado(self):
        """Doc 02 §27: stock anterior + compras - costo de venta."""
        sc = scope_stock(scope_op("OP-01"), "FAM-REP")
        o = val(self.values, "OPENING_STOCK", sc, "2027-05")
        p = val(self.values, "PURCHASES", sc, "2027-05")
        c = val(self.values, "COGS_FAMILY", sc, "2027-05")
        f = val(self.values, "CLOSING_STOCK", sc, "2027-05")
        self.assertEqual(f, o + p - c)

    def test_stock_encadena_periodos(self):
        sc = scope_stock(scope_op("OP-01"), "FAM-REP")
        cierre_abril = val(self.values, "CLOSING_STOCK", sc, "2027-04")
        inicio_mayo = val(self.values, "OPENING_STOCK", sc, "2027-05")
        self.assertEqual(cierre_abril, inicio_mayo)

    def test_cogs_por_familia_consolida_productos(self):
        """Doc 02 §27: productos -> familias -> costo de venta consolidado."""
        sc = scope_stock(scope_op("OP-01"), "FAM-REP")
        fam = val(self.values, "COGS_FAMILY", sc, "2027-01")
        unidad = self.version.configuration.unit("BU-01")
        productos = [p.id for p in unidad.products if p.family_id == "FAM-REP"]
        suma = sum(val(self.values, "COGS", f"OP:OP-01#P:{pid}", "2027-01")
                   for pid in productos)
        self.assertEqual(fam, suma)


class TestBalance(unittest.TestCase):
    def test_balance_que_no_cierra_bloquea(self):
        """Doc 02 §34: carga rechazada, no se incorpora parcialmente."""
        service, budget, version = bootstrap()
        for iv in version.inputs.values:
            if iv.concept is Concept.BALANCE_OPENING and iv.balance_item_id == "BI-CASH":
                iv.value = D(1)
        version.invalidate()
        findings = validate_balance(version.configuration, version.calculate(), "OPENING")
        self.assertTrue(findings)
        self.assertEqual(findings[0].code, "BALANCE_NOT_BALANCED")
        self.assertTrue(findings[0].blocking)

    def test_balance_correcto_no_genera_hallazgos(self):
        service, budget, version = bootstrap()
        self.assertEqual(validate_balance(version.configuration, version.calculate(), "OPENING"), [])

    def test_patrimonio_no_se_puede_cargar_si_es_calculado(self):
        """Doc 01 §26."""
        from app.domain.config import BalanceSource
        service, budget, version = bootstrap()
        cfg = version.configuration
        cfg.balance.items[-1].source = BalanceSource.CALCULATED
        iv = InputValue(concept=Concept.BALANCE_PROJECTED, value=D(1), balance_item_id="BI-RET")
        findings = validate_inputs(cfg, InputSet(values=[iv]))
        self.assertEqual(findings[0].code, "CALCULATED_VALUE_NOT_EDITABLE")


class TestRatiosAndAlerts(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.values = self.version.calculate()
        self.cfg = self.version.configuration

    def test_catalogo_completo(self):
        self.assertGreaterEqual(len(RATIO_CATALOG), 20)
        for code, r in RATIO_CATALOG.items():
            self.assertTrue(r.formula_text and r.metrics and r.required_inputs)

    def test_ebitda_margin(self):
        eb = val(self.values, "EBITDA", scope_co())
        s = val(self.values, "SALES", scope_co())
        r = val(self.values, "RATIO:EBITDA_MARGIN_PCT", scope_co())
        self.assertEqual(r, eb / s)

    def test_faltante_no_es_cero(self):
        """Doc 02 §41: un dato faltante no debe interpretarse como cero."""
        ratio = RATIO_CATALOG["STOCK_DAYS"]
        self.assertIsNone(ratio.compute({"STOCK_AVG": None, "COGS": D(100)}, 31))
        self.assertIsNone(ratio.compute({"STOCK_AVG": D(100), "COGS": D(0)}, 31))

    def test_objetivo_incumplido_genera_alerta_pero_no_bloquea(self):
        """Doc 02 §38/§40."""
        alerts = evaluate_objectives(self.cfg, self.values)
        codes = {a.code for a in alerts}
        self.assertIn("OBJECTIVE_NOT_MET", codes)
        report = self.service.validate_version(self.version)
        self.assertFalse(any(f.code == "OBJECTIVE_NOT_MET" for f in report["blocking"]))


class TestGovernance(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()

    def test_no_se_puede_cargar_antes_de_cerrar_configuracion(self):
        """Doc 01 §5."""
        service, budget, version = bootstrap(load_inputs=False, close_config=False)
        iv = InputValue(concept=Concept.SALES_QTY, period="2027-01", value=D(1),
                        operation_id="OP-01", product_id="P-001")
        with self.assertRaises(BudgetError) as ctx:
            service.submit_input("u.br01", version, iv, "budget.sales.load")
        self.assertEqual(ctx.exception.code, "CONFIGURATION_NOT_CLOSED")

    def test_configuracion_cerrada_bloquea_cambios_estructurales(self):
        """Doc 04 §10: 409 CONFIGURATION_LOCKED, no 400."""
        with self.assertRaises(BudgetError) as ctx:
            self.service.assert_configuration_open(self.version)
        self.assertEqual(ctx.exception.code, "CONFIGURATION_LOCKED")

    def test_scope_de_usuario(self):
        """Doc 04 §51: tener permiso no alcanza; hay que tener alcance."""
        iv = InputValue(concept=Concept.SALES_QTY, period="2027-01", value=D(10),
                        operation_id="OP-02", product_id="P-001")
        with self.assertRaises(BudgetError) as ctx:
            self.service.submit_input("u.br01", self.version, iv, "budget.sales.load")
        self.assertEqual(ctx.exception.code, "UNAUTHORIZED_SCOPE")

    def test_el_circuito_que_definio_el_cfo_se_respeta(self):
        """El workflow decía quién revisa y quién aprueba, y no lo miraba nadie."""
        balance = next(t for t in self.version.tasks.values() if t.concept == "BALANCE")
        balance.status = TaskStatus.IN_REVIEW
        with self.assertRaises(BudgetError) as ctx:
            self.service.reject_task("u.coo", self.version, balance.id, "no")
        self.assertEqual(ctx.exception.code, "UNAUTHORIZED_ROLE")
        ventas = next(t for t in self.version.tasks.values() if t.concept == "SALES")
        ventas.status = TaskStatus.IN_REVIEW
        self.service.reject_task("u.coo", self.version, ventas.id, "faltan datos")
        # queda RECHAZADA hasta que el que cargó la corrija: si volviera sola a
        # borrador, no habría manera de saber que la devolvieron
        self.assertEqual(ventas.status, TaskStatus.REJECTED)
        self.assertEqual(ventas.rejection, "faltan datos")

    def test_borrar_pide_el_mismo_permiso_que_cargar(self):
        """Borrar es cargar al revés: Finanzas borraba dotación ajena."""
        with self.assertRaises(BudgetError) as ctx:
            self.service.remove_inputs("u.fin", self.version, "MOV-01")
        self.assertEqual(ctx.exception.code, "UNAUTHORIZED")
        tarea = self.version.task_for("PAYROLL_SALARY", "CO")
        tarea.status = TaskStatus.APPROVED
        self.assertEqual(self.service.remove_inputs("u.br01", self.version, "MOV-01"), 2)
        self.assertEqual(tarea.status, TaskStatus.IN_REVIEW)

    def test_capacidad_requerida(self):
        iv = InputValue(concept=Concept.SALES_QTY, period="2027-01", value=D(10),
                        operation_id="OP-01", product_id="P-001")
        with self.assertRaises(BudgetError) as ctx:
            self.service.submit_input("u.admin", self.version, iv, "budget.sales.load")
        self.assertEqual(ctx.exception.code, "UNAUTHORIZED")

    def test_version_aprobada_es_inmutable(self):
        """Doc 04 §44/§63."""
        for t in self.version.tasks.values():
            t.status = TaskStatus.APPROVED
        self.version.status = VersionStatus.APPROVED
        with self.assertRaises(BudgetError) as ctx:
            self.service.change_fx_rate("u.cfo", self.version, "UYU", date(2027, 1, 15), D("0.026"))
        self.assertEqual(ctx.exception.code, "VERSION_IMMUTABLE")

    def test_nueva_version_clona_y_conserva_historico(self):
        """Doc 01 §9."""
        self.version.status = VersionStatus.APPROVED
        v2 = self.service.create_version("u.cfo", self.budget, self.version.id)
        self.assertEqual(v2.number, 2)
        self.assertEqual(len(v2.inputs.values), len(self.version.inputs.values))
        v2.inputs.values[0].value = D(1)
        self.assertNotEqual(self.version.inputs.values[0].value, D(1))
        self.assertIs(self.budget.versions[self.version.id], self.version)

    def test_vigente_requiere_aprobada(self):
        with self.assertRaises(BudgetError) as ctx:
            self.service.set_current("u.cfo", self.budget, self.version.id)
        self.assertEqual(ctx.exception.code, "VERSION_NOT_APPROVED")

    def test_rechazo_exige_motivo(self):
        """Doc 02 §48."""
        task = next(t for t in self.version.tasks.values() if t.concept == "SALES")
        task.status = TaskStatus.IN_REVIEW
        with self.assertRaises(BudgetError):
            self.service.reject_task("u.cfo", self.version, task.id, "")

    def test_aprobacion_parcial_sobrevive(self):
        """Doc 04 §47: lo aprobado que no fue afectado sigue aprobado."""
        t_sales_br1 = next(t for t in self.version.tasks.values()
                           if t.concept == "SALES" and t.scope_key == "OP:OP-01")
        t_sales_br3 = next(t for t in self.version.tasks.values()
                           if t.concept == "SALES" and t.scope_key == "OP:OP-03")
        for t in (t_sales_br1, t_sales_br3):
            t.status = TaskStatus.APPROVED
        iv = InputValue(concept=Concept.SALES_QTY, period="2027-01", value=D(2600),
                        operation_id="OP-01", product_id="P-001")
        self.service.submit_input("u.br01", self.version, iv, "budget.sales.load")
        self.assertEqual(t_sales_br1.status, TaskStatus.IN_REVIEW)
        self.assertEqual(t_sales_br3.status, TaskStatus.APPROVED)

    def test_auditoria_registra_antes_y_despues(self):
        """Doc 02 §55."""
        self.service.change_fx_rate("u.cfo", self.version, "UYU", date(2027, 1, 15), D("0.026"))
        ev = self.service.audit.filter(action="CHANGE_FX_RATE")[-1]
        self.assertEqual(ev.before, "0.025")
        self.assertEqual(ev.after, "0.026")
        self.assertEqual(ev.actor, "u.cfo")

    def test_no_se_puede_aprobar_con_tareas_pendientes(self):
        with self.assertRaises(BudgetError) as ctx:
            self.service.approve_version("u.cfo", self.version)
        self.assertIn(ctx.exception.code, ("PENDING_APPROVALS", "VERSION_NOT_READY"))


class TestResponsables(unittest.TestCase):
    """Cada centro de costo tiene quién carga sus valores, y Nómina los cubre a todos."""

    def test_el_perfil_del_centro_de_costo_es_quien_carga_sus_gastos(self):
        service, budget, version = bootstrap(load_inputs=False, close_config=False)
        cfg = version.configuration
        cfg.support_units[0].cost_centers[0].responsible_role = Role.FINANCE_AREA
        service.close_configuration("u.cfo", version)

        propia = next(t for t in version.tasks.values()
                      if t.concept == "EXPENSES" and t.scope_key == "CC:CC-01")
        self.assertEqual(propia.loader_role, Role.FINANCE_AREA)
        # la tarea general de gastos sigue el rol del workflow
        general = next(t for t in version.tasks.values()
                       if t.concept == "EXPENSES" and t.scope_key == "CO")
        self.assertEqual(general.loader_role, Role.ADMIN_AREA)

    def test_un_gasto_pertenece_a_la_tarea_de_su_centro_de_costo(self):
        service, budget, version = bootstrap()
        propia = next(t for t in version.tasks.values()
                      if t.concept == "EXPENSES" and t.scope_key == "CC:CC-01")
        general = next(t for t in version.tasks.values()
                       if t.concept == "EXPENSES" and t.scope_key == "CO")
        self.assertIs(version.task_for("EXPENSES", "CC:CC-01"), propia)
        # y por eso no queda además dentro de la general
        de_la_general = {iv.scope_key for iv in version.inputs_of(general)}
        self.assertNotIn("CC:CC-01", de_la_general)
        self.assertTrue(version.inputs_of(propia))

    def test_nomina_carga_todos_los_centros_de_costo(self):
        """Doc 01 §16: los centros salen de la estructura, no de una lista aparte."""
        from app.web.forms import build_form

        service, budget, version = bootstrap()
        cfg = version.configuration
        task = next(t for t in version.tasks.values() if t.concept == "PAYROLL_SALARY")
        self.assertEqual(task.loader_role, Role.PAYROLL_AREA)
        spec = build_form(version, task)
        etiquetas = [r.label for r in spec.rows]
        # primero la foto inicial de cada centro de costo...
        self.assertEqual(etiquetas[:len(cfg.cost_centers())],
                         [cc.name for cc, _k, _l in cfg.cost_centers()])
        # ...y después cada solicitud que hicieron las áreas, para valorizar
        self.assertIn("Alta — dos vendedores", etiquetas)
        self.assertIn("Ajuste — ascenso a jefatura", etiquetas)
        self.assertEqual(len(cfg.cost_centers()), len(cfg.operations) + 1)

    def test_agregar_una_operacion_agrega_su_centro_a_nomina(self):
        from app.domain.config import CostCenter, Operation
        from app.web.forms import build_form

        service, budget, version = bootstrap(load_inputs=False, close_config=False)
        cfg = version.configuration
        cfg.operations.append(Operation(
            id="OP-99", business_unit_id="BU-02", branch_id="BR-02",
            cost_center=CostCenter(id="CC-199", name="Servicios Salto")))
        service.close_configuration("u.cfo", version)
        task = next(t for t in version.tasks.values() if t.concept == "PAYROLL_SALARY")
        self.assertIn("Servicios Salto", [r.label for r in build_form(version, task).rows])
        # y su propia tarea de solicitudes, del responsable que le pusieron
        self.assertTrue(any(t.concept == "PAYROLL_HEADCOUNT" and t.scope_key == "CC:CC-199"
                            for t in version.tasks.values()))


class TestCicloDeVida(unittest.TestCase):
    """Lo que se rompió al revisar: el presupuesto no se podía cerrar."""

    def setUp(self):
        self.service, self.budget, self.version = bootstrap()

    def test_cargar_un_dato_pone_en_marcha_su_tarea(self):
        """Antes lo hacía la ruta web, así que una tarea cargada por la API o
        por la pantalla de dotación se quedaba muerta en NOT_STARTED."""
        task = next(t for t in self.version.tasks.values()
                    if t.concept == "SALES" and t.scope_key == "OP:OP-01")
        task.status = TaskStatus.NOT_STARTED
        self.service.submit_input("u.br01", self.version, InputValue(
            concept=Concept.SALES_QTY, period="2027-01", value=D(9),
            operation_id="OP-01", product_id="P-001"), "budget.sales.load")
        self.assertEqual(task.status, TaskStatus.DRAFT)
        self.service.submit_task("u.br01", self.version, task.id)
        self.assertEqual(self.service.approve_task("u.cfo", self.version, task.id).status,
                         TaskStatus.APPROVED)

    def test_una_tarea_sin_nada_que_cargar_se_puede_enviar(self):
        """Que un centro de costo no pida gente es una respuesta, no una tarea
        a medio hacer."""
        task = next(t for t in self.version.tasks.values()
                    if t.concept == "PAYROLL_HEADCOUNT")
        task.status = TaskStatus.NOT_STARTED
        self.service.submit_task("u.br01", self.version, task.id)
        self.assertEqual(task.status, TaskStatus.IN_REVIEW)

    def test_cerrar_la_configuracion_dos_veces_no_duplica_tareas(self):
        cuantas = len(self.version.tasks)
        with self.assertRaises(BudgetError) as ctx:
            self.service.close_configuration("u.cfo", self.version)
        self.assertEqual(ctx.exception.code, "CONFIGURATION_LOCKED")
        self.assertEqual(len(self.version.tasks), cuantas)

    def test_guardar_lo_mismo_no_desaprueba_nada(self):
        """Reenviar un formulario sin tocar nada no puede tirar abajo trabajo
        ya aprobado."""
        task = next(t for t in self.version.tasks.values()
                    if t.concept == "SALES" and t.scope_key == "OP:OP-01")
        task.status = TaskStatus.APPROVED
        igual = next(iv for iv in self.version.inputs.values
                     if iv.concept is Concept.SALES_QTY and iv.operation_id == "OP-01")
        self.service.submit_input("u.br01", self.version, igual.model_copy(deep=True),
                                  "budget.sales.load")
        self.assertEqual(task.status, TaskStatus.APPROVED)
        distinto = igual.model_copy(deep=True)
        distinto.value = igual.value + D(1)
        self.service.submit_input("u.br01", self.version, distinto, "budget.sales.load")
        self.assertEqual(task.status, TaskStatus.IN_REVIEW)


class TestLaPlataNoDesaparece(unittest.TestCase):
    """Lo peor que puede hacer un presupuesto es dar un número mal sin avisar."""

    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.cfg = self.version.configuration

    def test_la_asignacion_cierra_aunque_una_sucursal_no_venda(self):
        """Si el driver es cero el pool se repartía entre nadie y desaparecía."""
        self.version.inputs.values = [
            iv for iv in self.version.inputs.values
            if not (iv.operation_id == "OP-03" and iv.concept is Concept.SALES_AMOUNT)]
        self.version.invalidate()
        values = self.version.calculate()
        suma = sum(val(values, "RESULT_AFTER_ALLOCATION", scope_op(o.id))
                   for o in self.cfg.operations)
        self.assertAlmostEqual(float(suma), float(val(values, "EBITDA", scope_co())), places=4)

    def test_el_capex_de_una_sucursal_llega_al_total(self):
        antes = val(self.version.calculate(), "CAPEX", scope_co())
        self.version.inputs.add(InputValue(
            concept=Concept.CAPEX_AMOUNT, period="2027-03", value=D(100_000), currency="USD",
            branch_id="BR-01", capex_category_id="CAT-01"))
        self.version.inputs.add(InputValue(
            concept=Concept.CAPEX_AMOUNT, period="2027-04", value=D(70_000), currency="USD",
            cost_center_id="CC-101", capex_category_id="CAT-01"))
        self.version.invalidate()
        self.assertEqual(val(self.version.calculate(), "CAPEX", scope_co()), antes + D(170_000))

    def test_lo_que_el_motor_no_lee_se_rechaza_al_cargar(self):
        """Cargar contra un ámbito que nadie mira devolvía OK y la plata no
        entraba a ningún total."""
        casos = [
            InputValue(concept=Concept.EXPENSE_AMOUNT, period="2027-01", value=D(999_999),
                       currency="USD", expense_id="EXP-03", branch_id="BR-01"),
            InputValue(concept=Concept.COMMISSION_AMOUNT, period="2027-01", value=D(250_000),
                       currency="USD", operation_id="OP-03"),
            InputValue(concept=Concept.OPENING_STOCK, value=D(50_000), currency="USD",
                       business_unit_id="BU-01", family_id="FAM-REP"),
        ]
        for iv in casos:
            with self.assertRaises(BudgetError):
                self.service.submit_input("u.cfo", self.version, iv, "budget.expense.load")

    def test_la_nomina_respeta_la_vigencia_de_la_operacion(self):
        """Salto abre en junio: en enero no tiene ventas y tampoco tenía que
        tener nómina."""
        values = self.version.calculate()
        self.assertEqual(val(values, "PAYROLL", scope_op("OP-02"), "2027-01"), D(0))
        self.assertGreater(val(values, "PAYROLL", scope_op("OP-02"), "2027-06"), D(0))

    def test_los_saldos_de_apertura_usan_el_tipo_de_cambio_de_su_fecha(self):
        """Un saldo se valúa el día que existe, no el último día del ejercicio."""
        service, budget, version = bootstrap()
        cfg = version.configuration
        cfg.balance.currency = "UYU"
        version.fx.add_flat("UYU", date(2027, 1, 1), date(2027, 12, 31), D("0.03"))
        version.fx.add("UYU", cfg.fiscal_year.opening_balance_date, D("0.05"))
        version.invalidate()
        activo = version.calculate()[nk("ASSETS", scope_co(), "OPENING")]
        cargado = sum(iv.value for iv in version.inputs.values
                      if iv.concept is Concept.BALANCE_OPENING
                      and cfg.balance.items[[i.id for i in cfg.balance.items].index(
                          iv.balance_item_id)].section.value == "ASSET")
        self.assertEqual(activo, cargado * D("0.05"))


class TestVersiones(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        for t in self.version.tasks.values():
            t.status = TaskStatus.APPROVED
        self.service.approve_version("u.cfo", self.version)

    def test_la_version_nueva_reabre_la_configuracion(self):
        """Es la salida que promete CONFIGURATION_LOCKED. Sin esto, cerrada la
        primera versión el modelo no se podía cambiar nunca más."""
        v2 = self.service.create_version("u.cfo", self.budget, self.version.id)
        self.assertEqual(v2.configuration.status, ConfigStatus.DRAFT)
        self.assertEqual(v2.tasks, {})
        self.assertEqual(len(v2.inputs.values), len(self.version.inputs.values))
        v2.configuration.business_units[0].name = "Otro nombre"
        self.service.close_configuration("u.cfo", v2)
        self.assertEqual(v2.configuration.status, ConfigStatus.LOCKED)
        self.assertTrue(v2.tasks)
        self.assertEqual(self.version.configuration.business_units[0].name, "Repuestos")

    def test_el_tipo_de_cambio_no_se_comparte_entre_versiones(self):
        """Tocar el TC en la versión nueva movía los números de una versión ya
        aprobada: era la única cosa que no se copiaba en profundidad."""
        antes = self.version.calculate()[nk("SALES", scope_co(), FY)]
        v2 = self.service.create_version("u.cfo", self.budget, self.version.id)
        self.service.change_fx_rate("u.cfo", v2, "UYU", date(2027, 6, 15), D("0.05"))
        self.version.invalidate()
        self.assertEqual(self.version.calculate()[nk("SALES", scope_co(), FY)], antes)

    def test_no_se_aprueba_una_version_con_la_configuracion_abierta(self):
        v2 = self.service.create_version("u.cfo", self.budget, self.version.id)
        with self.assertRaises(BudgetError) as ctx:
            self.service.approve_version("u.cfo", v2)
        self.assertEqual(ctx.exception.code, "CONFIGURATION_NOT_CLOSED")


class TestDependencyGraph(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.values = self.version.calculate()

    def test_impacto_de_un_cambio(self):
        """Doc 04 §48: qué se afecta si cambia un input."""
        key = nk("SALES", "OP:OP-01#P:P-001", "2027-03")
        impact = self.version.impact_of([key])
        self.assertGreater(impact["affected_count"], 20)
        self.assertIn("EBITDA", impact["by_metric"])
        self.assertIn("RATIO:EBITDA_MARGIN_PCT", impact["by_metric"])

    def test_recalculo_incremental_da_lo_mismo_que_completo(self):
        """Doc 03 §43: recalcular sólo lo afectado no puede cambiar el resultado."""
        iv = InputValue(concept=Concept.SALES_QTY, period="2027-03", value=D(3000),
                        operation_id="OP-01", product_id="P-001")
        self.service.submit_input("u.br01", self.version, iv, "budget.sales.load")
        completo = self.version.calculate(force=True)
        self.version._values = self.values  # simula estado previo
        incremental = self.version.recalculate_from(
            [nk("SALES", "OP:OP-01#P:P-001", "2027-03")]
        )
        for m in ("SALES", "COGS", "EBITDA", "RESULT_AFTER_ALLOCATION"):
            self.assertEqual(completo[nk(m, scope_co(), FY)], incremental[nk(m, scope_co(), FY)],
                             f"difieren en {m}")

    def test_explicacion_de_un_valor(self):
        tree = self.version.graph.explain(nk("EBITDA", scope_co(), "2027-01"),
                                          self.values, depth=1)
        deps = {d["key"].split("|")[0] for d in tree["depends_on"]}
        self.assertEqual(deps, {"GROSS_MARGIN", "EXPENSES", "PAYROLL"})

    def test_sin_ciclos(self):
        order = self.version.graph.topological_order()
        self.assertEqual(len(order), len(self.version.graph.nodes))


class TestScenarios(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap()
        self.base = self.version.calculate()

    def test_escenario_no_modifica_la_base(self):
        """Doc 02 §51 / doc 03 §44."""
        sc = Scenario(id="SC1", name="Optimista", version_id=self.version.id,
                      adjustments=[ScenarioAdjustment(concept="SALES",
                                                      variation_type="PERCENTAGE",
                                                      variation=D("0.10"))])
        sim = run_scenario(self.version, sc)
        self.assertAlmostEqual(
            float(sim[nk("SALES", scope_co(), FY)]),
            float(self.base[nk("SALES", scope_co(), FY)]) * 1.10, places=2)
        self.assertEqual(self.version.calculate()[nk("SALES", scope_co(), FY)],
                         self.base[nk("SALES", scope_co(), FY)])

    def test_escenario_por_unidad(self):
        sc = Scenario(id="SC2", name="Sólo Repuestos", version_id=self.version.id,
                      adjustments=[ScenarioAdjustment(concept="SALES", variation_type="PERCENTAGE",
                                                      variation=D("0.20"),
                                                      business_unit_id="BU-01")])
        sim = run_scenario(self.version, sc)
        self.assertAlmostEqual(float(sim[nk("SALES", scope_bu("BU-01"), FY)]),
                               float(self.base[nk("SALES", scope_bu("BU-01"), FY)]) * 1.20, places=2)
        self.assertEqual(sim[nk("SALES", scope_bu("BU-02"), FY)],
                         self.base[nk("SALES", scope_bu("BU-02"), FY)])

    def test_escenario_de_costos_actua_sobre_el_supuesto(self):
        """No se puede tocar un calculado: subir costos = mover el margen."""
        sc = Scenario(id="SC3", name="Costos +5%", version_id=self.version.id,
                      adjustments=[ScenarioAdjustment(concept="COST", variation_type="PERCENTAGE",
                                                      variation=D("0.05"))])
        sim = run_scenario(self.version, sc)
        self.assertAlmostEqual(float(sim[nk("COGS", scope_co(), FY)]),
                               float(self.base[nk("COGS", scope_co(), FY)]) * 1.05, places=2)
        self.assertLess(sim[nk("EBITDA", scope_co(), FY)], self.base[nk("EBITDA", scope_co(), FY)])

    def test_no_se_puede_simular_sobre_un_calculado(self):
        sc = Scenario(id="SC4", name="EBITDA +1M", version_id=self.version.id,
                      adjustments=[ScenarioAdjustment(concept="EBITDA",
                                                      variation_type="ABSOLUTE",
                                                      variation=D(1_000_000))])
        with self.assertRaises(BudgetError) as ctx:
            run_scenario(self.version, sc)
        self.assertEqual(ctx.exception.code, "INVALID_SCENARIO_CONCEPT")

    def test_comparativa(self):
        sc = Scenario(id="SC5", name="Ventas +10%", version_id=self.version.id,
                      adjustments=[ScenarioAdjustment(concept="SALES", variation_type="PERCENTAGE",
                                                      variation=D("0.10"))])
        run_scenario(self.version, sc)
        rows = {r["metric"]: r for r in compare(self.version, sc)}
        self.assertGreater(rows["EBITDA"]["delta"], 0)


class TestImport(unittest.TestCase):
    def setUp(self):
        self.service, self.budget, self.version = bootstrap(load_inputs=False)

    def test_plantilla_se_genera_desde_la_configuracion(self):
        """Doc 02 §44: la planilla contiene únicamente lo que le corresponde al usuario."""
        from io import BytesIO
        from openpyxl import load_workbook
        from app.services.import_export import sales_template
        data = sales_template(self.version, "OP-02")
        ws = load_workbook(BytesIO(data))["Carga"]
        periodos = {ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)}
        self.assertTrue(all(p >= "2027-04" for p in periodos))  # abre en junio

    def test_importacion_atomica(self):
        """Doc 02 §45: un error rechaza toda la planilla."""
        from io import BytesIO
        from openpyxl import load_workbook
        from app.services.import_export import parse_sales_import, sales_template

        data = sales_template(self.version, "OP-01")
        wb = load_workbook(BytesIO(data))
        ws = wb["Carga"]
        col = [c.value for c in ws[1]].index("VALOR") + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).value = 10
        ws.cell(row=3, column=col).value = "diez"     # error en una fila
        buf = BytesIO(); wb.save(buf)

        result, parsed = parse_sales_import(self.version, buf.getvalue(), "OP-01", "u.br01")
        self.assertEqual(result.status, "REJECTED")
        self.assertEqual(parsed, [])
        self.assertEqual(result.errors[0].row, 3)
        self.assertIn("número", result.errors[0].error)

    def test_cero_valido_vacio_error(self):
        """Doc 02 §46: 0 es válido; vacío es error."""
        from io import BytesIO
        from openpyxl import load_workbook
        from app.services.import_export import parse_sales_import, sales_template

        data = sales_template(self.version, "OP-01")
        wb = load_workbook(BytesIO(data)); ws = wb["Carga"]
        col = [c.value for c in ws[1]].index("VALOR") + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).value = 0
        buf = BytesIO(); wb.save(buf)
        result, parsed = parse_sales_import(self.version, buf.getvalue(), "OP-01", "u.br01")
        self.assertEqual(result.status, "COMMITTED")
        self.assertTrue(parsed)

        wb = load_workbook(BytesIO(data)); ws = wb["Carga"]
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=col).value = 0
        ws.cell(row=2, column=col).value = None
        buf = BytesIO(); wb.save(buf)
        result, _ = parse_sales_import(self.version, buf.getvalue(), "OP-01", "u.br01")
        self.assertEqual(result.status, "REJECTED")
        self.assertIn("vacía", result.errors[0].error)


class TestCompleteness(unittest.TestCase):
    def test_faltantes_se_detectan(self):
        """Doc 02 §42: lo que se configura y es necesario debe cargarse."""
        service, budget, version = bootstrap(load_inputs=False)
        findings = missing_required_inputs(version.configuration, version.inputs)
        self.assertTrue(findings)
        self.assertTrue(all(f.blocking for f in findings))

    def test_demo_completa_no_tiene_faltantes(self):
        service, budget, version = bootstrap()
        findings = missing_required_inputs(version.configuration, version.inputs)
        self.assertEqual([f.message for f in findings], [])

    def test_ciclo_completo_hasta_version_vigente(self):
        """Doc 02 §64: criterio de aceptación global."""
        service, budget, version = bootstrap()
        report = service.validate_version(version)
        self.assertEqual([f.message for f in report["blocking"]], [])
        for t in list(version.tasks.values()):
            t.status = TaskStatus.APPROVED
        service.approve_version("u.cfo", version)
        service.set_current("u.cfo", budget, version.id)
        self.assertEqual(budget.current_version_id, version.id)
        self.assertEqual(version.status, VersionStatus.APPROVED)
        v2 = service.create_version("u.cfo", budget, version.id)
        self.assertTrue(v2.mutable)
        self.assertFalse(version.mutable)


if __name__ == "__main__":
    unittest.main()
